import operator
import openpyxl as pyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.xml.constants import MIN_COLUMN

def update_workbook(file_name):

    operations = {
        "multiplying" : operator.mul,
        "dividing" : operator.truediv,
        "adding" : operator.add,
        "subtracting" : operator.sub
    }


    wrkbk = pyxl.load_workbook(file_name)   # Load the user specified excel workbook.
    user_sheet = input("Which sheet number are you accessing: ")
    user_sheet_formatted = "Sheet" + user_sheet
    sheet = wrkbk[user_sheet_formatted]     # Retrieve the user specified desired excel sheet.
    print("Your selected sheet has", sheet.max_row, "rows.")
    start_row = int(input("Which row would you like to start from? Make sure to skip title/header rows: "))
    end_row = int(input("Which row would you like to end on: "))
    user_operation = input("What operation are we performing? Please type multiplying, dividing, adding, or subtracting: ").lower()

    while (user_operation not in ["multiplying", "dividing", "adding", "subtracting"]):
        user_operation = input("Invalid operation! Please try again: ")


    operation_value = float(input("Enter numerical value for the operation: "))
    col_selected = int(input(f"Which column number values should the program be {user_operation} {operation_value} by (this column should contain your initial/original values): "))
    new_value_col = int(input("Which column number should the program store the newly calculated values: "))

    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row, col_selected)
        try:
            corrected_price = operations[f"{user_operation}"](cell.value, operation_value)      # Allows the earlier, user specified, operation to be used here to calculate new price.
        except TypeError:
            corrected_price = ""
        corrected_price_cell = sheet.cell(row, new_value_col)   # Creates cell object that will hold newly calculated value.
        corrected_price_cell.value = corrected_price            # Value of said cell object now holds corrected price.


    print("""Would you like to a basic bar graph of the newly calculated values to be made? 
    *Note: If in the range of selected rows any of the selected rows had empty cells, the bar graph will also graph those empty cells resulting in a blank column for that row.*""")
    make_graph = input("Please type Yes or 'No': ").lower()
    if make_graph == "yes":
        graph_cell = input("In what cell would you like to place the graph in: ")
        values = Reference(sheet, min_row = start_row, max_row = end_row, min_col = new_value_col, max_col = new_value_col)     # Reference, to select a range of values - setup our potential future graph.
        chart = BarChart()
        chart.add_data(values)
        sheet.add_chart(chart, f"{graph_cell}")     # Place chart starting in user-defined cell.


    make_new_file = input("Would you like to save as a new file? Please type yes or no: ").lower()
    if make_new_file == "yes":
        new_file_name = input("What would you like to save the file as: ")
        wrkbk.save(new_file_name)
    else:
        wrkbk.save(file_name)
